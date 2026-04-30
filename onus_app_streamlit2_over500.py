import openpyxl
from collections import defaultdict
from datetime import datetime
from openpyxl.utils.datetime import from_excel

import streamlit as st
import pandas as pd

# ===========================
#       CONSTANTS
# ===========================
BONUS_OVER_400 = 20
BONUS_OVER_700 = 10
BONUS_AVG_130  = 20
BONUS_AVG_140  = 25
BONUS_AVG_150  = 35

QUANTITY_COL_NAME = "כמות"

REQUIRED_COLS = ["מספר חשבונית", "מוכרן", "תאריך", "מחיר נטו ליחידה"]


# ===========================
#       HELPERS
# ===========================
def parse_date(date_str: str):
    for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def find_header_row(sheet, required_cols):
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        col_idx = {}
        for i, value in enumerate(row):
            if value is None:
                continue
            key = str(value).strip()
            if key:
                col_idx[key] = i

        if col_idx and all(col in col_idx for col in required_cols):
            return row_idx, col_idx

    raise KeyError("לא נמצאה שורת כותרות מתאימה בקובץ (עמודות חובה לא אותרו יחד).")


def calculate_transaction_bonuses(sales_data, transactions_count):
    bonuses = defaultdict(float)
    details = defaultdict(lambda: defaultdict(float))

    for emp, daily in sales_data.items():
        for date, totals in daily.items():
            cnt = transactions_count[emp][date]
            avg = sum(totals) / cnt if cnt else 0

            for total in totals:
                if total > 400:
                    bonuses[emp] += BONUS_OVER_400
                    details[emp]["בונוס על עסקאות מעל 400"] += BONUS_OVER_400
                if total > 700:
                    bonuses[emp] += BONUS_OVER_700
                    details[emp]["בונוס על עסקאות מעל 700"] += BONUS_OVER_700

            if avg > 150:
                bonuses[emp] += BONUS_AVG_150
                details[emp]["בונוס על ממוצע עסקה"] += BONUS_AVG_150
            elif avg > 140:
                bonuses[emp] += BONUS_AVG_140
                details[emp]["בונוס על ממוצע עסקה"] += BONUS_AVG_140
            elif avg > 130:
                bonuses[emp] += BONUS_AVG_130
                details[emp]["בונוס על ממוצע עסקה"] += BONUS_AVG_130

    return bonuses, details


# ===========================
#       CORE LOGIC
# ===========================
def analyze_workbook(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = wb.active

    header_row, col_idx = find_header_row(sheet, REQUIRED_COLS)

    if QUANTITY_COL_NAME not in col_idx:
        col_idx[QUANTITY_COL_NAME] = None

    per_invoice        = defaultdict(lambda: defaultdict(list))
    transactions_count = defaultdict(lambda: defaultdict(int))
    daily_totals       = defaultdict(lambda: defaultdict(list))
    all_transactions   = []  # list of dicts: {מוכרן, תאריך, סכום עסקה}

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        invoice  = row[col_idx["מספר חשבונית"]]
        emp      = row[col_idx["מוכרן"]]
        date_raw = row[col_idx["תאריך"]]
        unit     = row[col_idx["מחיר נטו ליחידה"]]

        if col_idx[QUANTITY_COL_NAME] is not None:
            qty = row[col_idx[QUANTITY_COL_NAME]] or 1
        else:
            qty = 1

        if isinstance(date_raw, datetime):
            date = date_raw.date()
        elif isinstance(date_raw, (int, float)):
            date = from_excel(date_raw).date()
        elif isinstance(date_raw, str):
            date = parse_date(date_raw)
        else:
            date = None
        if date is None:
            continue

        if emp and unit is not None:
            line_total = unit * qty
            per_invoice[emp][(date, invoice)].append(line_total)

    for emp, invoices in per_invoice.items():
        for (date, _inv), lines in invoices.items():
            total = sum(lines)
            transactions_count[emp][date] += 1
            daily_totals[emp][date].append(total)
            all_transactions.append({
                "מוכרן": emp,
                "תאריך": date,
                "סכום עסקה": total,
            })

    bonuses, details = calculate_transaction_bonuses(
        daily_totals, transactions_count
    )

    return bonuses, details, daily_totals, transactions_count, all_transactions


def build_report_text(bonuses, details, daily_totals, transactions_count) -> str:
    lines = []

    lines.append("סיכום בונוסים כללי")
    lines.append("=" * 50)
    for emp, b in bonuses.items():
        lines.append(f"{emp}: {b:.2f} ₪")
    lines.append("")

    for emp, det in details.items():
        lines.append(emp)
        lines.append("-" * 50)
        for cat, amt in det.items():
            lines.append(f"{cat}: {amt:.2f} ₪")
        lines.append("")

    for emp, dates in daily_totals.items():
        lines.append(emp)
        lines.append("=" * 50)
        lines.append(f"{'תאריך':<10}\t{'ממוצע עסקה':<15}\t{'סך מכירות'}")
        lines.append("=" * 50)
        for date, totals in sorted(dates.items()):
            total = sum(totals)
            cnt   = transactions_count[emp][date]
            avg   = total / cnt if cnt else 0
            lines.append(
                f"{date.strftime('%d.%m'):<10}\t{avg:>12.2f}\t\t{total:>10.2f}"
            )
        lines.append("")

    return "\n".join(lines)


def build_summary_df(bonuses, daily_totals, transactions_count):
    rows = []
    for emp, dates in daily_totals.items():
        total_sales = sum(sum(totals) for totals in dates.values())
        total_tx    = sum(transactions_count[emp][d] for d in dates.keys())
        avg_tx      = total_sales / total_tx if total_tx else 0
        rows.append({
            "מוכרן": emp,
            "סך בונוס": bonuses.get(emp, 0.0),
            "סך מכירות": total_sales,
            "מספר עסקאות": total_tx,
            "ממוצע לעסקה": avg_tx,
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.sort_values("סך בונוס", ascending=False)
    return df


def build_employee_daily_df(emp, daily_totals, transactions_count):
    if emp not in daily_totals:
        return pd.DataFrame()

    rows = []
    for date, totals in sorted(daily_totals[emp].items()):
        day_total = sum(totals)
        cnt       = transactions_count[emp][date]
        avg       = day_total / cnt if cnt else 0
        rows.append({
            "תאריך": date.strftime("%d.%m.%Y"),
            "סך מכירות יומי": day_total,
            "מספר עסקאות": cnt,
            "ממוצע עסקה ביום": avg,
        })

    df = pd.DataFrame(rows)
    df = df.sort_values("תאריך", ascending=False)
    return df


# ===========================
#       STREAMLIT APP
# ===========================
def main():
    st.set_page_config(
        page_title="מנתח קובצי מכירה",
        page_icon="📊",
        layout="wide",
    )

    st.markdown("""
        <style>
            html, body, [class*="css"]  {
                direction: rtl !important;
                text-align: right !important;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Alef", sans-serif;
            }
            .block-container {
                padding-top: 2.5rem;
                padding-bottom: 2rem;
            }
            h1, h2, h3, h4 {
                text-align: right !important;
            }
            .stButton>button, .stDownloadButton>button {
                border-radius: 999px;
                padding: 0.5rem 1.5rem;
                font-weight: 600;
            }
            .stDataFrame, .stTable {
                direction: rtl !important;
                text-align: right !important;
            }
            table {
                direction: rtl !important;
            }
        </style>
    """, unsafe_allow_html=True)

    st.title("📊 מנתח קובצי מכירה - פילטר דינאמי")
    st.caption("העלה קובץ אקסל מהקופה והמערכת תחשב בונוסים ותציג דשבורד מסודר.")

    uploaded_file = st.file_uploader("בחר קובץ אקסל", type=["xlsx"])

    if uploaded_file is None:
        st.info("⬅️ כדי להתחיל, העלה קובץ אקסל בפורמט שאתה מקבל מהקופה.")
        return

    try:
        bonuses, details, daily_totals, transactions_count, all_transactions = analyze_workbook(
            uploaded_file
        )

        report_text = build_report_text(
            bonuses, details, daily_totals, transactions_count
        )

        summary_df  = build_summary_df(bonuses, daily_totals, transactions_count)
        all_tx_df   = pd.DataFrame(all_transactions)

        total_emps  = len(bonuses)
        total_bonus = sum(bonuses.values())
        total_sales = summary_df["סך מכירות"].sum() if not summary_df.empty else 0

        st.success("הקובץ עובד ונותח בהצלחה!")

        col1, col2, col3 = st.columns(3)
        col1.metric("מספר מוכרנים", f"{total_emps}")
        col2.metric("סך בונוסים", f"{total_bonus:,.0f} ₪")
        col3.metric("סך מכירות", f"{total_sales:,.0f} ₪")

        st.markdown("---")

        # ── טאבים: "כולם" + טאב לכל מוכרן ────────────────────
        emp_list   = list(bonuses.keys())
        tab_labels = ["כולם"] + emp_list
        tabs       = st.tabs(tab_labels)

        # ── טאב "כולם" ─────────────────────────────────────────
        with tabs[0]:
            st.subheader("סיכום לפי מוכרן")
            if summary_df.empty:
                st.warning("לא נמצאו נתונים להצגה.")
            else:
                st.dataframe(
                    summary_df.style.format({
                        "סך בונוס": "{:,.0f} ₪",
                        "סך מכירות": "{:,.0f} ₪",
                        "מספר עסקאות": "{:,.0f}",
                        "ממוצע לעסקה": "{:,.1f} ₪",
                    }),
                    use_container_width=True,
                )

            st.markdown("---")
            st.subheader("🔍 פילטר עסקאות לפי סכום")

            min_price_all = st.number_input(
                "מכירות מעל:",
                min_value=0,
                value=0,
                step=50,
                key="price_filter_all"
            )

            filtered_all = all_tx_df[all_tx_df["סכום עסקה"] > min_price_all]

            if not filtered_all.empty:
                counts_df = (
                    filtered_all.groupby("מוכרן")
                    .size()
                    .reset_index(name="כמות עסקאות")
                    .sort_values("כמות עסקאות", ascending=False)
                )
                st.caption(f"סה״כ **{len(filtered_all)}** עסקאות מעל {min_price_all:,} ₪")
                st.dataframe(counts_df, use_container_width=True)
            else:
                st.info("לא נמצאו עסקאות מעל הסכום שנבחר.")

        # ── טאב לכל מוכרן ──────────────────────────────────────
        for i, emp in enumerate(emp_list):
            with tabs[i + 1]:

                # ביצועים יומיים + פירוט בונוסים
                emp_daily_df = build_employee_daily_df(emp, daily_totals, transactions_count)

                col_emp1, col_emp2 = st.columns([2, 1])

                with col_emp1:
                    st.markdown(f"#### ביצועים יומיים – {emp}")
                    if emp_daily_df.empty:
                        st.write("אין נתונים להצגה עבור מוכרן זה.")
                    else:
                        st.dataframe(
                            emp_daily_df.style.format({
                                "סך מכירות יומי": "{:,.0f} ₪",
                                "מספר עסקאות": "{:,.0f}",
                                "ממוצע עסקה ביום": "{:,.1f} ₪",
                            }),
                            use_container_width=True,
                            height=400,
                        )

                with col_emp2:
                    st.markdown("#### פירוט בונוסים")
                    emp_det = details.get(emp, {})
                    if not emp_det:
                        st.write("אין בונוסים מפורטים למוכרן זה.")
                    else:
                        for cat, amt in emp_det.items():
                            st.write(f"• **{cat}** – {amt:,.0f} ₪")

                # ── פילטר עסקאות בודדות ────────────────────────
                st.markdown("---")
                st.subheader("🔍 פילטר עסקאות לפי סכום")

                min_price_emp = st.number_input(
                    "מכירות מעל:",
                    min_value=0,
                    value=0,
                    step=50,
                    key=f"price_filter_{emp}"
                )

                emp_tx = all_tx_df[
                    (all_tx_df["מוכרן"] == emp) &
                    (all_tx_df["סכום עסקה"] > min_price_emp)
                ].copy()

                if emp_tx.empty:
                    st.info("לא נמצאו עסקאות מעל הסכום שנבחר.")
                else:
                    st.caption(f"נמצאו **{len(emp_tx)}** עסקאות מעל {min_price_emp:,} ₪")
                    emp_tx["תאריך"] = emp_tx["תאריך"].apply(
                        lambda d: d.strftime("%d.%m.%Y")
                    )
                    emp_tx = emp_tx[["תאריך", "סכום עסקה"]].sort_values("תאריך", ascending=False)
                    st.dataframe(
                        emp_tx.style.format({"סכום עסקה": "{:,.0f} ₪"}),
                        use_container_width=True,
                        height=350,
                    )

        # ── דוח טקסט ───────────────────────────────────────────
        st.markdown("---")
        with st.expander("📄 דוח טקסט מלא (לצפייה / העתקה)", expanded=False):
            st.text(report_text)

        st.download_button(
            label="⬇️ הורד דוח כקובץ TXT",
            data=report_text.encode("utf-8"),
            file_name="bonus_report.txt",
            mime="text/plain",
        )

    except KeyError as e:
        st.error(f"שגיאה בעמודות הקובץ: {e}")
    except Exception as e:
        st.error(f"נכשל עיבוד הקובץ: {e}")


if __name__ == "__main__":
    main()

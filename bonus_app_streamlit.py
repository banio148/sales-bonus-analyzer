import openpyxl
from collections import defaultdict
from datetime import datetime
from openpyxl.utils.datetime import from_excel
import streamlit as st

# ===========================
#       CONSTANTS
# ===========================
BONUS_OVER_400 = 20
BONUS_OVER_700 = 10
BONUS_AVG_130  = 20
BONUS_AVG_140  = 25
BONUS_AVG_150  = 35

QUANTITY_COL_NAME = "כמות"

REQUIRED_COLS = ["מספר חשבונית", "מוכרן", "תאריך", "מחיר נטו ליחידה"]


# ===========================
#       HELPERS
# ===========================
def parse_date(date_str: str):
    """נסיון לפענח תאריך ממחרוזת בכמה פורמטים."""
    for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def find_header_row(sheet, required_cols):
    """
    מחפשת את שורת הכותרות בגיליון.
    מדלגת על שורות ריקות / שורה עם השם שלך בלבד וכו'.
    מחזירה: (מספר_שורה, dict של שם עמודה -> אינדקס).
    """
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        col_idx = {}
        for i, value in enumerate(row):
            if value is None:
                continue
            key = str(value).strip()
            if key:
                col_idx[key] = i

        if col_idx and all(col in col_idx for col in required_cols):
            return row_idx, col_idx

    raise KeyError("לא נמצאה שורת כותרות מתאימה בקובץ (עמודות חובה לא אותרו יחד).")


def calculate_transaction_bonuses(sales_data, transactions_count):
    bonuses = defaultdict(float)
    details = defaultdict(lambda: defaultdict(float))

    for emp, daily in sales_data.items():
        for date, totals in daily.items():
            cnt = transactions_count[emp][date]
            avg = sum(totals) / cnt if cnt else 0

            for total in totals:
                if total > 400:
                    bonuses[emp] += BONUS_OVER_400
                    details[emp]["בונוס על עסקאות מעל 400"] += BONUS_OVER_400
                if total > 700:
                    bonuses[emp] += BONUS_OVER_700
                    details[emp]["בונוס על עסקאות מעל 700"] += BONUS_OVER_700

            if avg > 150:
                bonuses[emp] += BONUS_AVG_150
                details[emp]["בונוס על ממוצע עסקה"] += BONUS_AVG_150
            elif avg > 140:
                bonuses[emp] += BONUS_AVG_140
                details[emp]["בונוס על ממוצע עסקה"] += BONUS_AVG_140
            elif avg > 130:
                bonuses[emp] += BONUS_AVG_130
                details[emp]["בונוס על ממוצע עסקה"] += BONUS_AVG_130

    return bonuses, details


# ===========================
#       CORE LOGIC
# ===========================
def analyze_workbook(file_obj):
    """
    מקבל קובץ אקסל (file-like) ומחזיר:
    bonuses, details, daily_totals, transactions_count
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = wb.active

    # למצוא שורת כותרות אמיתית (גם אם יש 5–6 שורות ריקות / שורה עם השם שלך)
    header_row, col_idx = find_header_row(sheet, REQUIRED_COLS)

    # אם אין עמודת כמות – נניח תמיד 1
    if QUANTITY_COL_NAME not in col_idx:
        col_idx[QUANTITY_COL_NAME] = None

    per_invoice        = defaultdict(lambda: defaultdict(list))
    transactions_count = defaultdict(lambda: defaultdict(int))
    daily_totals       = defaultdict(lambda: defaultdict(list))

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        invoice = row[col_idx["מספר חשבונית"]]
        emp     = row[col_idx["מוכרן"]]
        date_raw= row[col_idx["תאריך"]]
        unit    = row[col_idx["מחיר נטו ליחידה"]]

        # כמות
        if col_idx[QUANTITY_COL_NAME] is not None:
            qty = row[col_idx[QUANTITY_COL_NAME]] or 1
        else:
            qty = 1

        # תאריך
        if isinstance(date_raw, datetime):
            date = date_raw.date()
        elif isinstance(date_raw, (int, float)):
            date = from_excel(date_raw).date()
        elif isinstance(date_raw, str):
            date = parse_date(date_raw)
        else:
            date = None
        if date is None:
            continue

        if emp and unit is not None:
            line_total = unit * qty
            per_invoice[emp][(date, invoice)].append(line_total)

    # סכימה לפי חשבונית -> יום -> מוכרן
    for emp, invoices in per_invoice.items():
        for (date, _inv), lines in invoices.items():
            total = sum(lines)
            transactions_count[emp][date] += 1
            daily_totals[emp][date].append(total)

    bonuses, details = calculate_transaction_bonuses(
        daily_totals, transactions_count
    )

    return bonuses, details, daily_totals, transactions_count


def build_report_text(bonuses, details, daily_totals, transactions_count) -> str:
    lines = []

    lines.append("סיכום בונוסים כללי")
    lines.append("=" * 50)
    for emp, b in bonuses.items():
        lines.append(f"{emp}: {b:.2f} ₪")
    lines.append("")

    for emp, det in details.items():
        lines.append(emp)
        lines.append("-" * 50)
        for cat, amt in det.items():
            lines.append(f"{cat}: {amt:.2f} ₪")
        lines.append("")

    for emp, dates in daily_totals.items():
        lines.append(emp)
        lines.append("=" * 50)
        lines.append(f"{'תאריך':<10}\t{'ממוצע עסקה':<15}\t{'סך מכירות'}")
        lines.append("=" * 50)
        for date, totals in sorted(dates.items()):
            total = sum(totals)
            cnt   = transactions_count[emp][date]
            avg   = total / cnt if cnt else 0
            lines.append(
                f"{date.strftime('%d.%m'):<10}\t{avg:>12.2f}\t\t{total:>10.2f}"
            )
        lines.append("")

    return "\n".join(lines)


# ===========================
#       STREAMLIT APP
# ===========================
def main():
    st.set_page_config(page_title="מנתח קובצי מכירה", layout="wide")

    st.markdown("""
    <style>
        html, body, [class*="css"]  {
            direction: rtl !important;
            text-align: right !important;
            font-family: "Alef", sans-serif;
        }

        /* יישור לימין של תיבת הטקסט של הדוח */
        .stText, .stMarkdown, .stTextInput, pre, code {
            direction: rtl !important;
            text-align: right !important;
        }

        /* יישור לימין של כפתורים, כותרות וכל הרכיבים */
        .stButton, .stDownloadButton, .stSelectbox, .stFileUploader {
            direction: rtl !important;
            text-align: right !important;
        }

        /* יישור לטבלאות (אם תוסיף בהמשך) */
        table {
            direction: rtl !important;
            text-align: right !important;
        }
    </style>
""", unsafe_allow_html=True)

    st.title("מנתח קובצי מכירה – חישוב בונוסים")
    st.write("העלה קובץ אקסל כפי שמתקבל מהקופה, והמערכת תחפש לבד את שורת הכותרות ותחשב בונוסים.")

    uploaded_file = st.file_uploader("בחר קובץ אקסל", type=["xlsx"])

    if uploaded_file is not None:
        try:
            bonuses, details, daily_totals, transactions_count = analyze_workbook(
                uploaded_file
            )
            report_text = build_report_text(
                bonuses, details, daily_totals, transactions_count
            )

            st.success("הקובץ עובד ונותח בהצלחה!")

            st.subheader("תצוגת דוח")
            st.text(report_text)

            st.download_button(
                label="הורד דוח כקובץ TXT",
                data=report_text.encode("utf-8"),
                file_name="bonus_report.txt",
                mime="text/plain",
            )

        except KeyError as e:
            st.error(f"שגיאה בעמודות הקובץ: {e}")
        except Exception as e:
            st.error(f"נכשל עיבוד הקובץ: {e}")


if __name__ == "__main__":
    main()


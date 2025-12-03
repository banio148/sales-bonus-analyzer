import openpyxl
from collections import defaultdict
from datetime import datetime
from openpyxl.utils.datetime import from_excel

import streamlit as st
import pandas as pd

# ===========================
#       CONSTANTS
# ===========================
BONUS_OVER_400 = 20
BONUS_OVER_700 = 10
BONUS_AVG_130  = 20
BONUS_AVG_140  = 25
BONUS_AVG_150  = 35

QUANTITY_COL_NAME = "×›××•×ª"

REQUIRED_COLS = ["××¡×¤×¨ ×—×©×‘×•× ×™×ª", "××•×›×¨×Ÿ", "×ª××¨×™×š", "××—×™×¨ × ×˜×• ×œ×™×—×™×“×”"]


# ===========================
#       HELPERS
# ===========================
def parse_date(date_str: str):
    """× ×¡×™×•×Ÿ ×œ×¤×¢× ×— ×ª××¨×™×š ×××—×¨×•×–×ª ×‘×›××” ×¤×•×¨××˜×™×."""
    for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def find_header_row(sheet, required_cols):
    """
    ××—×¤×©×ª ××ª ×©×•×¨×ª ×”×›×•×ª×¨×•×ª ×‘×’×™×œ×™×•×Ÿ.
    ××“×œ×’×ª ×¢×œ ×©×•×¨×•×ª ×¨×™×§×•×ª / ×©×•×¨×” ×¢× ×”×©× ×©×œ×š ×‘×œ×‘×“ ×•×›×•'.
    ××—×–×™×¨×”: (××¡×¤×¨_×©×•×¨×”, dict ×©×œ ×©× ×¢××•×“×” -> ××™× ×“×§×¡).
    """
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        col_idx = {}
        for i, value in enumerate(row):
            if value is None:
                continue
            key = str(value).strip()
            if key:
                col_idx[key] = i

        if col_idx and all(col in col_idx for col in required_cols):
            return row_idx, col_idx

    raise KeyError("×œ× × ××¦××” ×©×•×¨×ª ×›×•×ª×¨×•×ª ××ª××™××” ×‘×§×•×‘×¥ (×¢××•×“×•×ª ×—×•×‘×” ×œ× ××•×ª×¨×• ×™×—×“).")


def calculate_transaction_bonuses(sales_data, transactions_count):
    bonuses = defaultdict(float)
    details = defaultdict(lambda: defaultdict(float))

    for emp, daily in sales_data.items():
        for date, totals in daily.items():
            cnt = transactions_count[emp][date]
            avg = sum(totals) / cnt if cnt else 0

            for total in totals:
                if total > 400:
                    bonuses[emp] += BONUS_OVER_400
                    details[emp]["×‘×•× ×•×¡ ×¢×œ ×¢×¡×§××•×ª ××¢×œ 400"] += BONUS_OVER_400
                if total > 700:
                    bonuses[emp] += BONUS_OVER_700
                    details[emp]["×‘×•× ×•×¡ ×¢×œ ×¢×¡×§××•×ª ××¢×œ 700"] += BONUS_OVER_700

            if avg > 150:
                bonuses[emp] += BONUS_AVG_150
                details[emp]["×‘×•× ×•×¡ ×¢×œ ×××•×¦×¢ ×¢×¡×§×”"] += BONUS_AVG_150
            elif avg > 140:
                bonuses[emp] += BONUS_AVG_140
                details[emp]["×‘×•× ×•×¡ ×¢×œ ×××•×¦×¢ ×¢×¡×§×”"] += BONUS_AVG_140
            elif avg > 130:
                bonuses[emp] += BONUS_AVG_130
                details[emp]["×‘×•× ×•×¡ ×¢×œ ×××•×¦×¢ ×¢×¡×§×”"] += BONUS_AVG_130

    return bonuses, details


# ===========================
#       CORE LOGIC
# ===========================
def analyze_workbook(file_obj):
    """
    ××§×‘×œ ×§×•×‘×¥ ××§×¡×œ (file-like) ×•××—×–×™×¨:
    bonuses, details, daily_totals, transactions_count
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = wb.active

    # ×œ××¦×•× ×©×•×¨×ª ×›×•×ª×¨×•×ª ×××™×ª×™×ª (×’× ×× ×™×© 5â€“6 ×©×•×¨×•×ª ×¨×™×§×•×ª / ×©×•×¨×” ×¢× ×”×©× ×©×œ×š)
    header_row, col_idx = find_header_row(sheet, REQUIRED_COLS)

    # ×× ××™×Ÿ ×¢××•×“×ª ×›××•×ª â€“ × × ×™×— ×ª××™×“ 1
    if QUANTITY_COL_NAME not in col_idx:
        col_idx[QUANTITY_COL_NAME] = None

    per_invoice        = defaultdict(lambda: defaultdict(list))
    transactions_count = defaultdict(lambda: defaultdict(int))
    daily_totals       = defaultdict(lambda: defaultdict(list))

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        invoice = row[col_idx["××¡×¤×¨ ×—×©×‘×•× ×™×ª"]]
        emp     = row[col_idx["××•×›×¨×Ÿ"]]
        date_raw= row[col_idx["×ª××¨×™×š"]]
        unit    = row[col_idx["××—×™×¨ × ×˜×• ×œ×™×—×™×“×”"]]

        # ×›××•×ª
        if col_idx[QUANTITY_COL_NAME] is not None:
            qty = row[col_idx[QUANTITY_COL_NAME]] or 1
        else:
            qty = 1

        # ×ª××¨×™×š
        if isinstance(date_raw, datetime):
            date = date_raw.date()
        elif isinstance(date_raw, (int, float)):
            date = from_excel(date_raw).date()
        elif isinstance(date_raw, str):
            date = parse_date(date_raw)
        else:
            date = None
        if date is None:
            continue

        if emp and unit is not None:
            line_total = unit * qty
            per_invoice[emp][(date, invoice)].append(line_total)

    # ×¡×›×™××” ×œ×¤×™ ×—×©×‘×•× ×™×ª -> ×™×•× -> ××•×›×¨×Ÿ
    for emp, invoices in per_invoice.items():
        for (date, _inv), lines in invoices.items():
            total = sum(lines)
            transactions_count[emp][date] += 1
            daily_totals[emp][date].append(total)

    bonuses, details = calculate_transaction_bonuses(
        daily_totals, transactions_count
    )

    return bonuses, details, daily_totals, transactions_count


def build_report_text(bonuses, details, daily_totals, transactions_count) -> str:
    lines = []

    lines.append("×¡×™×›×•× ×‘×•× ×•×¡×™× ×›×œ×œ×™")
    lines.append("=" * 50)
    for emp, b in bonuses.items():
        lines.append(f"{emp}: {b:.2f} â‚ª")
    lines.append("")

    for emp, det in details.items():
        lines.append(emp)
        lines.append("-" * 50)
        for cat, amt in det.items():
            lines.append(f"{cat}: {amt:.2f} â‚ª")
        lines.append("")

    for emp, dates in daily_totals.items():
        lines.append(emp)
        lines.append("=" * 50)
        lines.append(f"{'×ª××¨×™×š':<10}\t{'×××•×¦×¢ ×¢×¡×§×”':<15}\t{'×¡×š ××›×™×¨×•×ª'}")
        lines.append("=" * 50)
        for date, totals in sorted(dates.items()):
            total = sum(totals)
            cnt   = transactions_count[emp][date]
            avg   = total / cnt if cnt else 0
            lines.append(
                f"{date.strftime('%d.%m'):<10}\t{avg:>12.2f}\t\t{total:>10.2f}"
            )
        lines.append("")

    return "\n".join(lines)


def build_summary_df(bonuses, daily_totals, transactions_count):
    rows = []
    for emp, dates in daily_totals.items():
        total_sales = sum(sum(totals) for totals in dates.values())
        total_tx    = sum(transactions_count[emp][d] for d in dates.keys())
        avg_tx      = total_sales / total_tx if total_tx else 0
        rows.append({
            "××•×›×¨×Ÿ": emp,
            "×¡×š ×‘×•× ×•×¡": bonuses.get(emp, 0.0),
            "×¡×š ××›×™×¨×•×ª": total_sales,
            "××¡×¤×¨ ×¢×¡×§××•×ª": total_tx,
            "×××•×¦×¢ ×œ×¢×¡×§×”": avg_tx,
        })

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.sort_values("×¡×š ×‘×•× ×•×¡", ascending=False)
    return df


def build_employee_daily_df(emp, daily_totals, transactions_count):
    if emp not in daily_totals:
        return pd.DataFrame()

    rows = []
    for date, totals in sorted(daily_totals[emp].items()):
        day_total = sum(totals)
        cnt       = transactions_count[emp][date]
        avg       = day_total / cnt if cnt else 0
        rows.append({
            "×ª××¨×™×š": date.strftime("%d.%m.%Y"),
            "×¡×š ××›×™×¨×•×ª ×™×•××™": day_total,
            "××¡×¤×¨ ×¢×¡×§××•×ª": cnt,
            "×××•×¦×¢ ×¢×¡×§×” ×‘×™×•×": avg,
        })

    df = pd.DataFrame(rows)
    df = df.sort_values("×ª××¨×™×š", ascending=False)
    return df


# ===========================
#       STREAMLIT APP
# ===========================
def main():
    st.set_page_config(
        page_title="×× ×ª×— ×§×•×‘×¦×™ ××›×™×¨×”",
        page_icon="ğŸ“Š",
        layout="wide",
    )

    # RTL + ×™×™×©×•×¨ ×œ×™××™×Ÿ + ×˜××¥' ××•×“×¨× ×™
    st.markdown("""
        <style>
            html, body, [class*="css"]  {
                direction: rtl !important;
                text-align: right !important;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Alef", sans-serif;
            }

            .block-container {
                padding-top: 2.5rem;
                padding-bottom: 2rem;
            }

            h1, h2, h3, h4 {
                text-align: right !important;
            }

            .stButton>button, .stDownloadButton>button {
                border-radius: 999px;
                padding: 0.5rem 1.5rem;
                font-weight: 600;
            }

            /* ×˜×‘×œ××•×ª */
            .stDataFrame, .stTable {
                direction: rtl !important;
                text-align: right !important;
            }

            table {
                direction: rtl !important;
            }
        </style>
    """, unsafe_allow_html=True)

    st.title("ğŸ“Š ×× ×ª×— ×§×•×‘×¦×™ ××›×™×¨×” â€“ ×—×™×©×•×‘ ×‘×•× ×•×¡×™×")
    st.caption("×”×¢×œ×” ×§×•×‘×¥ ××§×¡×œ ××”×§×•×¤×” ×•×”××¢×¨×›×ª ×ª×—×©×‘ ×‘×•× ×•×¡×™× ×•×ª×¦×™×’ ×“×©×‘×•×¨×“ ××¡×•×“×¨.")

    uploaded_file = st.file_uploader("×‘×—×¨ ×§×•×‘×¥ ××§×¡×œ", type=["xlsx"])

    if uploaded_file is None:
        st.info("â¬…ï¸ ×›×“×™ ×œ×”×ª×—×™×œ, ×”×¢×œ×” ×§×•×‘×¥ ××§×¡×œ ×‘×¤×•×¨××˜ ×©××ª×” ××§×‘×œ ××”×§×•×¤×”.")
        return

    try:
        bonuses, details, daily_totals, transactions_count = analyze_workbook(
            uploaded_file
        )

        # ×“×•×— ×˜×§×¡×˜ ××œ×
        report_text = build_report_text(
            bonuses, details, daily_totals, transactions_count
        )

        summary_df = build_summary_df(bonuses, daily_totals, transactions_count)

        # ×›×¨×˜×™×¡×™ ×¡×˜×˜×•×¡ ×¢×œ×™×•× ×™×
        total_emps   = len(bonuses)
        total_bonus  = sum(bonuses.values())
        total_sales  = summary_df["×¡×š ××›×™×¨×•×ª"].sum() if not summary_df.empty else 0

        st.success("×”×§×•×‘×¥ ×¢×•×‘×“ ×•× ×•×ª×— ×‘×”×¦×œ×—×”!")

        col1, col2, col3 = st.columns(3)
        col1.metric("××¡×¤×¨ ××•×›×¨× ×™×", f"{total_emps}")
        col2.metric("×¡×š ×‘×•× ×•×¡×™×", f"{total_bonus:,.0f} â‚ª")
        col3.metric("×¡×š ××›×™×¨×•×ª", f"{total_sales:,.0f} â‚ª")

        st.markdown("---")

        # ×˜×‘×œ×ª ×¡×™×›×•×
        st.subheader("×¡×™×›×•× ×œ×¤×™ ××•×›×¨×Ÿ")
        if summary_df.empty:
            st.warning("×œ× × ××¦××• × ×ª×•× ×™× ×œ×”×¦×’×”.")
        else:
            st.dataframe(
                summary_df.style.format({
                    "×¡×š ×‘×•× ×•×¡": "{:,.0f} â‚ª",
                    "×¡×š ××›×™×¨×•×ª": "{:,.0f} â‚ª",
                    "××¡×¤×¨ ×¢×¡×§××•×ª": "{:,.0f}",
                    "×××•×¦×¢ ×œ×¢×¡×§×”": "{:,.1f} â‚ª",
                }),
                use_container_width=True,
            )

        st.markdown("### ×¤×™×¨×•×˜ ×œ×¤×™ ××•×›×¨×Ÿ")

        if bonuses:
            emp_list = list(bonuses.keys())
            selected_emp = st.selectbox("×‘×—×¨ ××•×›×¨×Ÿ ×œ×¤×™×¨×•×˜:", emp_list)

            emp_daily_df = build_employee_daily_df(
                selected_emp, daily_totals, transactions_count
            )

            col_emp1, col_emp2 = st.columns([2, 1])

            with col_emp1:
                st.markdown(f"#### ×‘×™×¦×•×¢×™× ×™×•××™×™× â€“ {selected_emp}")
                if emp_daily_df.empty:
                    st.write("××™×Ÿ × ×ª×•× ×™× ×œ×”×¦×’×” ×¢×‘×•×¨ ××•×›×¨×Ÿ ×–×”.")
                else:
                    st.dataframe(
                        emp_daily_df.style.format({
                            "×¡×š ××›×™×¨×•×ª ×™×•××™": "{:,.0f} â‚ª",
                            "××¡×¤×¨ ×¢×¡×§××•×ª": "{:,.0f}",
                            "×××•×¦×¢ ×¢×¡×§×” ×‘×™×•×": "{:,.1f} â‚ª",
                        }),
                        use_container_width=True,
                        height=400,
                    )

            with col_emp2:
                st.markdown("#### ×¤×™×¨×•×˜ ×‘×•× ×•×¡×™×")
                emp_det = details.get(selected_emp, {})
                if not emp_det:
                    st.write("××™×Ÿ ×‘×•× ×•×¡×™× ××¤×•×¨×˜×™× ×œ××•×›×¨×Ÿ ×–×”.")
                else:
                    for cat, amt in emp_det.items():
                        st.write(f"â€¢ **{cat}** â€“ {amt:,.0f} â‚ª")

        # ×“×•×— ×˜×§×¡×˜ ×’×•×œ××™ ×œ-export / ×‘×“×™×§×•×ª
        with st.expander("ğŸ“„ ×“×•×— ×˜×§×¡×˜ ××œ× (×œ×¦×¤×™×™×” / ×”×¢×ª×§×”)", expanded=False):
            st.text(report_text)

        # ×›×¤×ª×•×¨ ×”×•×¨×“×”
        st.download_button(
            label="â¬‡ï¸ ×”×•×¨×“ ×“×•×— ×›×§×•×‘×¥ TXT",
            data=report_text.encode("utf-8"),
            file_name="bonus_report.txt",
            mime="text/plain",
        )

    except KeyError as e:
        st.error(f"×©×’×™××” ×‘×¢××•×“×•×ª ×”×§×•×‘×¥: {e}")
    except Exception as e:
        st.error(f"× ×›×©×œ ×¢×™×‘×•×“ ×”×§×•×‘×¥: {e}")


if __name__ == "__main__":
    main()
